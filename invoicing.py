# invoicing.py
from __future__ import annotations

from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional, Iterable
import json
import uuid
import os
import re
import csv
from datetime import date, datetime
import functools

# ---------- internal data paths (app-local) ----------
DATA_DIR = Path(__file__).resolve().parent / "data"
INVOICES_DIR = DATA_DIR / "invoices"  # internal storage (unchanged)
SETTINGS_PATH = DATA_DIR / "invoicing_settings.json"

# ---------- user-visible default ----------
DEFAULT_USER_INVOICE_ROOT = Path.home() / "Baymaxx Invoices"


# ---------- small utils ----------
def _ensure_dirs() -> None:
    """Make sure app-local data dirs exist (data/invoices)."""
    INVOICES_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def _atomic_write_text(path: Path, text: str) -> None:
    """Write via a temp file and replace to avoid partial writes."""
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(text, encoding="utf-8")
    os.replace(tmp, path)


def _new_id() -> str:
    return str(uuid.uuid4())


# ======================================================================
# Canonical line-item + totals helpers (SINGLE SOURCE OF TRUTH)
# ======================================================================

def _ensure_line_items(inv: dict) -> list:
    """Make sure inv['line_items'] exists and is a list; return it."""
    if not isinstance(inv.get("line_items"), list):
        inv["line_items"] = []
    return inv["line_items"]


def recompute_totals(inv: dict) -> None:
    """Recalculate invoice total from line items (qty * unit_price)."""
    subtotal = 0.0
    for li in inv.get("line_items", []):
        qty = float(li.get("qty", 0) or 0)
        price = float(li.get("unit_price", 0) or 0)
        amt = round(qty * price, 2)
        li["amount"] = amt
        subtotal += amt

    inv.setdefault("totals", {})
    inv["totals"]["subtotal"] = round(subtotal, 2)
    tax_rate = float(inv.get("tax_rate", 0.0))
    inv["totals"]["tax"] = round(subtotal * tax_rate, 2)
    inv["totals"]["total"] = round(inv["totals"]["subtotal"] + inv["totals"]["tax"], 2)


# canonical add_line_item
def add_line_item(inv: Dict[str, Any], description: str, qty: float, unit_price: float) -> None:
    """Append a line item and recalc totals."""
    items = _ensure_line_items(inv)
    qty = float(qty or 0)
    unit_price = float(unit_price or 0.0)
    amount = round(qty * unit_price, 2)

    items.append({
        "description": (description or "").strip(),
        "qty": qty,
        "unit_price": unit_price,
        "amount": amount,
    })
    recompute_totals(inv)


# compat aliases used elsewhere
def add_item(inv: dict, description: str, qty: float, unit_price: float) -> None:
    add_line_item(inv, description, qty, unit_price)

def _add_item(inv: dict, description: str, qty: float, unit_price: float) -> None:
    add_line_item(inv, description, qty, unit_price)

def _recompute_totals(inv: dict) -> None:
    recompute_totals(inv)


# ======================================================================
# Messages support & compat helpers
# ======================================================================

try:
    UNIT_PRICE_SMS
except NameError:
    UNIT_PRICE_SMS = 0.14


# ---------------- phone resolution (re-usable) ----------------
def _normalize_site_key(s: str) -> str:
    u = (s or '').upper().strip()
    for suf in (' VOICE', ' SMS'):
        if u.endswith(suf):
            u = u[:-len(suf)].strip()
    if '–' in u:
        u = u.split('–', 1)[0].strip()
    u = ' '.join(u.split())
    return u


# ---------------- CSV aggregation (month/year) ----------------
def _extract_row_datetime(row: dict, kind: str):
    """
    Try to parse a datetime from a CSV row for the given kind.
    """
    def _norm(h: str) -> str:
        return re.sub(r"[\s_\-]+", "", (h or "").strip().lower())

    if kind == "calls":
        wanted = {"starttime", "start", "calldate"}
    else:  # messages
        wanted = {"sentdate", "date", "messagedate", "senddate", "timestamp"}

    val = None
    for key, raw in row.items():
        nk = _norm(key)
        if any(tok in nk for tok in wanted) and raw not in (None, ""):
            val = str(raw).strip()
            break

    if not val:
        return None

    try:
        return datetime.fromisoformat(val.replace("Z", "+00:00"))
    except Exception:
        pass

    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", val)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d)
        except Exception:
            pass

    m = re.search(r"(\d{1,2})/(\d{1,2})/(\d{2,4})", val)
    if m:
        mo, d, y = map(int, m.groups())
        if y < 100:
            y += 2000
        try:
            return datetime(y, mo, d)
        except Exception:
            pass

    return None


def _aggregate_rows_by_site(files_with_sites, kind: str, year: int, month: int) -> dict[str, int]:
    """files_with_sites: List[Tuple[path, site_name_or_None]]"""
    from collections import defaultdict

    counts = defaultdict(int)
    for path, site_name in (files_with_sites or []):
        site = site_name or Path(path).stem
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    dt = _extract_row_datetime(row, kind)
                    if not dt:
                        continue
                    if dt.year == year and dt.month == month:
                        counts[site] += 1
        except Exception:
            pass
    return dict(counts)


# ---------- settings (remember user's chosen folder) ----------
def _load_settings() -> Dict[str, Any]:
    try:
        if SETTINGS_PATH.exists():
            return json.loads(SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def _save_settings(d: Dict[str, Any]) -> None:
    _ensure_dirs()
    _atomic_write_text(SETTINGS_PATH, json.dumps(d, indent=2, ensure_ascii=False) + "\n")


def get_remembered_invoice_root() -> Optional[Path]:
    s = _load_settings()
    p = s.get("invoice_root")
    if not p:
        return None
    pp = Path(p)
    return pp if pp.exists() and pp.is_dir() else None


def set_remembered_invoice_root(path: Path) -> None:
    s = _load_settings()
    s["invoice_root"] = str(path)
    _save_settings(s)


def ensure_invoice_root(parent: Optional[object] = None) -> Optional[Path]:
    remembered = get_remembered_invoice_root()
    if remembered:
        return remembered

    if DEFAULT_USER_INVOICE_ROOT.exists():
        set_remembered_invoice_root(DEFAULT_USER_INVOICE_ROOT)
        return DEFAULT_USER_INVOICE_ROOT

    askyesno = None
    askdirectory = None
    try:
        from tkinter import messagebox, filedialog
        askyesno = messagebox.askyesno
        askdirectory = filedialog.askdirectory
    except Exception:
        DEFAULT_USER_INVOICE_ROOT.mkdir(parents=True, exist_ok=True)
        set_remembered_invoice_root(DEFAULT_USER_INVOICE_ROOT)
        return DEFAULT_USER_INVOICE_ROOT

    create = askyesno(
        "Create Invoices Folder",
        f"Baymaxx needs a place to save invoices.\n\n"
        f"Create this folder?\n\n{DEFAULT_USER_INVOICE_ROOT}"
    )
    if create:
        try:
            DEFAULT_USER_INVOICE_ROOT.mkdir(parents=True, exist_ok=True)
            set_remembered_invoice_root(DEFAULT_USER_INVOICE_ROOT)
            return DEFAULT_USER_INVOICE_ROOT
        except Exception:
            pass

    chosen = askdirectory(
        title="Choose a folder to save invoices",
        initialdir=str(Path.home())
    )
    if not chosen:
        return None

    chosen_path = Path(chosen)
    try:
        chosen_path.mkdir(parents=True, exist_ok=True)
    except Exception:
        return None

    set_remembered_invoice_root(chosen_path)
    return chosen_path


def invoice_output_dir() -> Path:
    p = get_remembered_invoice_root()
    if p:
        return p
    DEFAULT_USER_INVOICE_ROOT.mkdir(parents=True, exist_ok=True)
    set_remembered_invoice_root(DEFAULT_USER_INVOICE_ROOT)
    return DEFAULT_USER_INVOICE_ROOT


def _ensure_out_dir_for_invoice(inv: Dict[str, Any], out_dir: str | Path | None) -> Path:
    """Resolve and create the output directory for a given invoice."""
    if out_dir is not None:
        out_path = Path(out_dir)
    else:
        root = invoice_output_dir() or INVOICES_DIR
        inv_type = str(inv.get("type") or "invoice").capitalize()
        today_str = date.today().strftime("%m-%d-%Y")
        folder_name = f"{inv_type} {today_str}"
        out_path = Path(root) / folder_name

    out_path.mkdir(parents=True, exist_ok=True)
    return out_path


# ---------- invoice API ----------
def new_monthly_invoice(
    year: int,
    month: int,
    client_id: str | None = None,
    client_name_snapshot: str | None = None,
    tax_rate: float = 0.0,
) -> Dict[str, Any]:
    _ensure_dirs()
    inv: Dict[str, Any] = {
        "id": _new_id(),
        "type": "monthly",
        "period": {"year": int(year), "month": int(month)},
        "client_id": client_id,
        "client_name_snapshot": client_name_snapshot,
        "created_at": None,
        "tax_rate": float(tax_rate),
        "line_items": [],
        "totals": {"subtotal": 0.0, "tax": 0.0, "total": 0.0},
        "notes": "",
    }
    return inv


def set_client(inv: Dict[str, Any], client_id: str, client_name_snapshot: str) -> None:
    inv["client_id"] = client_id
    inv["client_name_snapshot"] = client_name_snapshot


def set_tax_rate(inv: Dict[str, Any], tax_rate: float) -> None:
    inv["tax_rate"] = float(tax_rate)
    recompute_totals(inv)


def save_invoice(inv: Dict[str, Any]) -> Path:
    _ensure_dirs()
    if not inv.get("id"):
        inv["id"] = _new_id()
    recompute_totals(inv)
    path = INVOICES_DIR / f"{inv['id']}.json"
    _atomic_write_text(path, json.dumps(inv, indent=2, ensure_ascii=False) + "\n")
    return path


def load_invoice(invoice_id: str) -> Dict[str, Any] | None:
    path = INVOICES_DIR / f"{invoice_id}.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def delete_invoice(invoice_id: str) -> bool:
    path = INVOICES_DIR / f"{invoice_id}.json"
    if not path.exists():
        return False
    try:
        path.unlink()
        return True
    except Exception:
        return False


def list_invoices() -> List[Dict[str, Any]]:
    _ensure_dirs()
    out: List[Dict[str, Any]] = []
    for p in sorted(INVOICES_DIR.glob("*.json")):
        try:
            doc = json.loads(p.read_text(encoding="utf-8"))
            out.append({
                "id": doc.get("id"),
                "type": doc.get("type"),
                "period": doc.get("period"),
                "client_id": doc.get("client_id"),
                "client_name": doc.get("client_name_snapshot"),
                "total": (doc.get("totals") or {}).get("total", 0.0),
            })
        except Exception:
            continue
    return out


# ---------- CSV helpers (kind + source number) ----------
def _norm(s: str) -> str:
    return re.sub(r"[\s_\-]+", "", (s or "").strip().lower())


def _detect_kind(fieldnames: List[str]) -> str:
    """
    Decide 'messages' vs 'calls' using Twilio-ish headers.

    Heuristics:
      - Messages if we see NumSegments (or close variants) or other SMS/Messages markers.
      - Calls if we see Duration, Start/End Time, CallSid, AnsweredBy, etc.
      - If both appear, prefer calls when any duration-like header is present,
        otherwise prefer messages when NumSegments is present.
    """
    if not fieldnames:
        return "unknown"

    normalized = {_norm(h) for h in fieldnames}

    # message indicators
    msg_tokens = (
        "numsegments", "numofsegments", "sentdate", "messagedate", "smsstatus",
        "messagingservice", "message", "body"
    )
    has_msg = any(any(tok in h for tok in msg_tokens) for h in normalized)

    # call indicators
    call_tokens = (
        "duration", "starttime", "endtime", "calldate", "callsid", "answeredby",
        "callstatus", "call", "price"
    )
    has_call = any(any(tok in h for tok in call_tokens) for h in normalized)

    # strongest signals
    has_numsegments = any(("numsegments" in h) or ("numofsegments" in h) for h in normalized)
    has_duration = any("duration" in h for h in normalized)

    if has_duration:
        return "calls"
    if has_numsegments:
        return "messages"

    if has_call and not has_msg:
        return "calls"
    if has_msg and not has_call:
        return "messages"

    return "unknown"



def sniff_csv(path: str | Path) -> Tuple[str, List[str]]:
    p = Path(path)
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
    return _detect_kind(headers), headers


def _clean_phone(raw: str) -> str:
    if not raw:
        return ""
    raw = raw.strip()
    lead_plus = raw.startswith("+")
    digits = re.sub(r"\D+", "", raw)
    return ("+" if lead_plus else "") + digits


def identify_source(path: str | Path) -> Dict[str, Any]:
    p = Path(path)
    kind, headers = sniff_csv(p)

    normalized = [_norm(h) for h in headers]

    # Priority columns for extracting the Twilio/site number
    if kind == "calls":
        candidate_names = [
            "to", "called", "destination",   # often your Twilio number
            "from", "callerid", "caller",    # fallback
            "sender", "source"
        ]
    else:
        candidate_names = ["from", "sender", "source", "callerid", "caller"]

    header_index = None
    for i, hn in enumerate(normalized):
        if hn in candidate_names:
            header_index = i
            break

    raw_number = ""
    if header_index is not None:
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if header_index < len(row):
                    raw = row[header_index].strip()
                    if raw:
                        raw_number = raw
                        break

    return {
        "kind": kind,
        "raw_number": raw_number,
        "number": _clean_phone(raw_number),
        "headers": headers,
    }



# ---------- matching helpers (CSV -> site by last-4) ----------
def _digits_only(s: str) -> str:
    return re.sub(r"\D+", "", s or "")


def _match_site_by_last4(clients_doc, phone_digits: str) -> dict | None:
    if not phone_digits:
        return None
    last4 = phone_digits[-4:]

    candidates = clients_doc.get("clients") if isinstance(clients_doc, dict) else clients_doc
    if not isinstance(candidates, list):
        return None

    for c in candidates:
        client_name = (c or {}).get("name", "")
        for d in (c or {}).get("divisions", []) or []:
            division_name = (d or {}).get("name", "")
            for s in (d or {}).get("sites", []) or []:
                site_phone_digits = _digits_only((s or {}).get("phone", ""))
                if site_phone_digits.endswith(last4):
                    return {
                        "client_id": (c or {}).get("id"),
                        "client_name": client_name,
                        "division_id": (d or {}).get("id"),
                        "division_name": division_name,
                        "site_id": (s or {}).get("id"),
                        "site_name": (s or {}).get("name", ""),
                        "site_phone": site_phone_digits,
                        "matched_last4": last4,
                    }
    return None


def identify_csv_and_phone(path: str | Path, clients_doc=None) -> dict:
    base = identify_source(path)
    phone_digits = _digits_only(base.get("number") or base.get("raw_number") or "")
    match = _match_site_by_last4(clients_doc, phone_digits) if clients_doc else None

    return {
        "kind": base.get("kind", "unknown"),
        "phone": phone_digits,
        "match": match,
        "headers": base.get("headers", []),
    }


# ---------- month/year validation for CSVs ----------
def _ym_from_any_date(s: str) -> tuple[int | None, int | None]:
    if not s:
        return (None, None)
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if not m:
        return (None, None)
    y = int(m.group(1))
    mo = int(m.group(2))
    return (y, mo)


def check_csv_month_year(path: str | Path, kind: str, year: int, month: int) -> tuple[bool, dict]:
    p = Path(path)
    try:
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
            norm = [_norm(h) for h in headers]

            if kind == "messages":
                candidates = {"sentdate", "date", "timestamp"}
            elif kind == "calls":
                candidates = {"starttime", "start", "calldate"}
            else:
                return (False, {"in": 0, "out": 0, "rows": 0})

            idx = None
            for i, hn in enumerate(norm):
                if hn in candidates:
                    idx = i
                    break
            if idx is None:
                return (False, {"in": 0, "out": 0, "rows": 0})

            n_in = n_out = n_total = 0
            for row in reader:
                n_total += 1
                val = row[idx] if idx < len(row) else ""
                y, m = _ym_from_any_date(val)
                if y == year and m == month:
                    n_in += 1
                else:
                    n_out += 1

        return (n_out == 0 and n_total > 0, {"in": n_in, "out": n_out, "rows": n_total})
    except Exception:
        return (False, {"in": 0, "out": 0, "rows": 0})


# ======================================================================
# Kind + description helpers (canonical)
# ======================================================================

_KIND_TOKENS = ("VOICE", "SMS")

def _normalize_kind(kind: str | None) -> str | None:
    if not kind:
        return None
    k = str(kind).strip().upper()
    if k in _KIND_TOKENS:
        return k
    if "CALL" in k or "VOICE" in k:
        return "VOICE"
    if "MSG" in k or "SMS" in k or "TEXT" in k:
        return "SMS"
    return None

def make_site_description(site_name: str, kind: str | None) -> str:
    name = (site_name or "").strip()
    k = _normalize_kind(kind)

    upper_name = name.upper()
    if re.search(r"\b(VOICE|SMS)\b\s*$", upper_name):
        return name

    if k:
        return f"{name} {k}".strip()

    return name


# ---------- export naming helpers ----------
def monthly_filename_for_today() -> str:
    today = date.today()
    return f"{today:%Y-%m-%d} Monthly.json"


def monthly_output_path(root: Path | None = None) -> Path:
    root = (root or invoice_output_dir())
    root.mkdir(parents=True, exist_ok=True)
    return root / monthly_filename_for_today()


# ---------- Excel template → PDF export helpers ----------
@functools.lru_cache(maxsize=1)
def _load_clients_doc() -> dict:
    try:
        here = Path(__file__).resolve().parent
        clients_path = here / "data" / "clients.json"
        with clients_path.open("r", encoding="utf-8") as f:
            doc = json.load(f)
        return doc or {}
    except FileNotFoundError:
        return {}
    except Exception as e:
        print("WARNING: could not load clients.json for site ordering:", e)
        return {}

def _find_client_address(clients_doc: dict, name_snapshot: str | None) -> list[str]:
    if not name_snapshot:
        return []
    items = clients_doc.get("clients") or clients_doc.get("items") or []
    for c in items:
        if (c.get("name") or "").strip() == name_snapshot.strip():
            addr = (c.get("address") or "").strip()
            lines = [name_snapshot]
            if addr:
                for line in addr.splitlines():
                    if line.strip():
                        lines.append(line.strip())
            return lines[:3]
    return [name_snapshot]


# ---------- ordering helpers ----------
def _iter_sites_in_clients_order(clients_doc: dict):
    for client in (clients_doc.get("clients") or []):
        for div in (client.get("divisions") or []):
            for site in (div.get("sites") or []):
                name = (site.get("name") or "").strip()
                if name:
                    yield name


def _ordered_site_items(by_site: dict[str, int]) -> list[tuple[str, int]]:
    clients_doc = _load_clients_doc()

    ordered: list[tuple[str, int]] = []
    used: set[str] = set()

    for site_name in _iter_sites_in_clients_order(clients_doc):
        if site_name in by_site:
            ordered.append((site_name, by_site[site_name]))
            used.add(site_name)

    leftovers = sorted(k for k in by_site.keys() if k not in used)
    for site_name in leftovers:
        ordered.append((site_name, by_site[site_name]))

    return ordered


# ---------- phone matching / decorator helpers ----------
def _build_priority_phone_map(inv: dict) -> dict[str, str]:
    phones: dict[str, str] = {}

    # 1) from invoice data (highest priority)
    try:
        sp = inv.get("site_phones")
        if isinstance(sp, dict):
            for raw_name, v in sp.items():
                if not raw_name or not v:
                    continue
                last4 = str(v)[-4:]
                phones[raw_name] = last4
                try:
                    nk = _normalize_site_key(raw_name)
                    phones.setdefault(nk, last4)
                except Exception:
                    pass
    except Exception:
        pass

    # 2) from clients.json (fallback)
    try:
        clients_doc = _load_clients_doc()
        items = clients_doc.get("clients") or clients_doc.get("items") or []
        for c in items:
            for d in c.get("divisions", []) or []:
                for s in d.get("sites", []) or []:
                    raw_name = (s.get("name") or "").strip()
                    raw_phone = s.get("phone") or ""
                    digits = _digits_only(raw_phone)
                    if raw_name and digits:
                        last4 = digits[-4:]
                        phones.setdefault(raw_name, last4)
                        try:
                            nk = _normalize_site_key(raw_name)
                            phones.setdefault(nk, last4)
                        except Exception:
                            pass
    except Exception:
        pass

    return phones


def _lookup_last4(phones: dict[str, str], desc: str) -> str | None:
    if not desc:
        return None
    key = _normalize_site_key(desc)
    return phones.get(key)


# ======================================================================
# Voice invoice helpers
# ======================================================================

UNIT_PRICE_VOICE = 0.14  # USD per call (flat), per user spec

def _normalize_headers(headers: list[str]) -> list[str]:
    return [re.sub(r"[\s_\-]+", "", (h or "").strip().lower()) for h in headers]

def _date_col_index(headers: list[str], kind: str) -> int | None:
    """Return index of the date column we should check for this kind."""
    norm = lambda s: re.sub(r"[\s_\-]+", "", s.strip().lower())
    normed = [norm(h) for h in headers]
    if kind == "messages":
        targets = {"sentdate", "date", "timestamp"}
    elif kind == "calls":
        targets = {"starttime", "start", "calldate"}
    else:
        targets = set()

    for i, n in enumerate(normed):
        if any(t in n for t in targets):
            return i
    return None


_date_re = re.compile(r"(\d{4})-(\d{2})-(\d{2})")

def _ym_from_cell(cell: str) -> tuple[int | None, int | None]:
    if not cell:
        return (None, None)
    m = _date_re.search(cell)
    if not m:
        return (None, None)
    y, mo, _ = m.groups()
    try:
        return (int(y), int(mo))
    except Exception:
        return (None, None)

def count_rows_calls_csv(path: str | Path, filter_year: int | None = None, filter_month: int | None = None) -> int:
    p = Path(path)
    with p.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        if filter_year is None or filter_month is None:
            return sum(1 for _ in reader)
        ci = _date_col_index(headers, "calls")
        if ci is None:
            return 0
        n = 0
        for row in reader:
            cell = row[ci] if ci < len(row) else ""
            y, m = _ym_from_cell(cell)
            if y == filter_year and m == filter_month:
                n += 1
        return n

def build_voice_line_item(site_name: str | None, qty: int, unit_price: float = UNIT_PRICE_VOICE) -> dict:
    """
    Build a single VOICE line item.
    Always produce '<Site Name> VOICE' (unless site already ends with VOICE).
    """
    base = (site_name or "").strip()
    desc = make_site_description(base, "VOICE") if base else "VOICE"

    return {
        "description": desc,
        "qty": float(qty),
        "unit_price": float(unit_price),
    }


def aggregate_voice_items_from_csvs(files_with_sites, year=None, month=None):
    by_site: dict[str, int] = {}
    for csv_path, site_name in files_with_sites:
        qty = count_rows_calls_csv(csv_path, year, month) if (year and month) else count_rows_calls_csv(csv_path)
        label = site_name or Path(csv_path).stem
        by_site[label] = by_site.get(label, 0) + int(qty)

    items: list[dict[str, Any]] = []
    for site_key, qty in _ordered_site_items(by_site):
        items.append(build_voice_line_item(site_key or None, qty))
    return items


def add_voice_items_to_invoice(inv: Dict[str, Any],
                               files_with_sites: list[tuple[str | Path, str | None]],
                               year: int | None = None,
                               month: int | None = None,
                               unit_price: float = UNIT_PRICE_VOICE) -> Dict[str, Any]:
    """
    Aggregate VOICE items from CSVs and append them to inv.

    Important:
    - DO NOT strip the VOICE label.
    - DO NOT pre-append (-last4) here.
      The PDF/template exporter will call decorate_with_last4_kind()
      and add (-last4) while preserving VOICE.
    """
    items = aggregate_voice_items_from_csvs(files_with_sites, year, month)

    for it in items:
        raw_desc = str(it.get("description", "")).strip()

        # Ensure it has a VOICE label (but don't double-append)
        desc = make_site_description(raw_desc, "VOICE")

        add_line_item(inv, desc, it.get("qty", 0), unit_price)

    return inv


# ======================================================================
# Export helpers + PDF/CSV
# ======================================================================

def invoice_filename(inv: Dict[str, Any], ext: str) -> str:
    per = inv.get("period") or {}
    y = per.get("year")
    m = per.get("month")
    ym = f"{y}-{int(m):02d}" if y and m else "unknown"
    return f"Invoice-{ym}-{inv.get('id','')}.{ext.lstrip('.')}"


def export_invoice_csv(inv: Dict[str, Any], out_dir: str | Path | None = None) -> Path:
    recompute_totals(inv)
    out_dir = _ensure_out_dir_for_invoice(inv, out_dir)

    csv_path = out_dir / invoice_filename(inv, "csv")

    import io, csv as _csv
    buf = io.StringIO()
    w = _csv.writer(buf, lineterminator="\n")
    w.writerow(["Description", "Qty", "Unit Price", "Amount"])
    for li in inv.get("line_items", []):
        w.writerow([
            li.get("description", ""),
            li.get("qty", 0),
            li.get("unit_price", 0),
            li.get("amount", 0),
        ])
    w.writerow([])
    totals = inv.get("totals", {})
    w.writerow(["Subtotal", "", "", totals.get("subtotal", 0)])
    w.writerow(["Tax", "", "", totals.get("tax", 0)])
    w.writerow(["Total", "", "", totals.get("total", 0)])

    csv_path.write_text(buf.getvalue(), encoding="utf-8")
    return csv_path


# ---------- Simple PDF export (ReportLab) ----------
def export_invoice_pdf(inv: Dict[str, Any], out_dir: str | Path | None = None) -> Path:
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
    except Exception as e:
        raise RuntimeError("reportlab is required: pip install reportlab") from e

    recompute_totals(inv)
    out_dir = _ensure_out_dir_for_invoice(inv, out_dir)
    pdf_path = out_dir / invoice_filename(inv, "pdf")

    c = canvas.Canvas(str(pdf_path), pagesize=LETTER)
    width, height = LETTER
    x_margin = 0.75 * inch
    y = height - 0.75 * inch

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x_margin, y, "INVOICE")
    y -= 18

    per = inv.get("period") or {}
    period_text = f"Period: {per.get('year','')} - {int(per.get('month',0)):02d}" if per else ""
    c.setFont("Helvetica", 10)
    if period_text:
        c.drawString(x_margin, y, period_text)
        y -= 14
    inv_id = inv.get("id", "")
    c.drawString(x_margin, y, f"Invoice ID: {inv_id}")
    y -= 20

    snap = inv.get("client_name_snapshot", "")
    if snap:
        c.drawString(x_margin, y, f"Client: {snap}")
        y -= 16

    div_name = inv.get("division_name", "")
    if div_name:
        c.drawString(x_margin, y, f"Division: {div_name}")
        y -= 16

    c.setFont("Helvetica-Bold", 10)
    col_desc_x = x_margin
    col_qty_x  = x_margin + 4.6 * inch
    col_unit_x = x_margin + 5.4 * inch
    col_amt_x  = x_margin + 6.3 * inch
    c.drawString(col_desc_x, y, "Description")
    c.drawString(col_qty_x,  y, "Qty")
    c.drawString(col_unit_x, y, "Unit Price")
    c.drawString(col_amt_x,  y, "Amount")
    y -= 12
    c.line(x_margin, y, width - x_margin, y)
    y -= 8

    c.setFont("Helvetica", 10)
    for li in inv.get("line_items", []):
        if y < 1.3 * inch:
            c.showPage()
            y = height - 0.75 * inch
            c.setFont("Helvetica-Bold", 10)
            c.drawString(col_desc_x, y, "Description")
            c.drawString(col_qty_x,  y, "Qty")
            c.drawString(col_unit_x, y, "Unit Price")
            c.drawString(col_amt_x,  y, "Amount")
            y -= 12
            c.line(x_margin, y, width - x_margin, y)
            y -= 8
            c.setFont("Helvetica", 10)

        raw_desc = str(li.get("description", ""))
        try:
            desc = decorate_with_last4_kind(inv, raw_desc)
        except Exception:
            desc = raw_desc

        qty_val = li.get("qty", 0)
        try:
            qty = str(int(round(float(qty_val))))
        except Exception:
            qty = str(qty_val)

        unit = f"{li.get('unit_price', 0):.2f}"
        amt  = f"{li.get('amount', 0):.2f}"

        c.drawString(col_desc_x, y, desc)
        c.drawRightString(col_qty_x + 0.5 * inch, y, qty)
        c.drawRightString(col_unit_x + 0.8 * inch, y, unit)
        c.drawRightString(col_amt_x + 0.8 * inch, y, amt)
        y -= 14

    y -= 6
    c.line(x_margin, y, width - x_margin, y)
    y -= 12

    totals = inv.get("totals", {})
    c.setFont("Helvetica-Bold", 10)
    c.drawRightString(col_unit_x + 0.8 * inch, y, "Subtotal:")
    c.drawRightString(col_amt_x + 0.8 * inch, y, f"{totals.get('subtotal', 0):.2f}")
    y -= 14
    c.drawRightString(col_unit_x + 0.8 * inch, y, "Tax:")
    c.drawRightString(col_amt_x + 0.8 * inch, y, f"{totals.get('tax', 0):.2f}")
    y -= 14
    c.drawRightString(col_unit_x + 0.8 * inch, y, "Total:")
    c.drawRightString(col_amt_x + 0.8 * inch, y, f"{totals.get('total', 0):.2f}")

    c.showPage()
    c.save()
    return pdf_path


# ======================================================================
# Decoration helpers
# ======================================================================

import re as _re_dec

def _infer_kind_and_base(desc: str) -> tuple[str | None, str]:
    d = (desc or "").strip()
    up = d.upper()

    m = re.search(r"\b(VOICE|SMS)\b\s*$", up)
    if m:
        kind = m.group(1)
        base = re.sub(r"\b(VOICE|SMS)\b\s*$", "", up).strip()
        return kind, base

    has_voice = re.search(r"\bVOICE\b", up) is not None
    has_sms   = re.search(r"\bSMS\b", up) is not None

    if has_voice and not has_sms:
        kind = "VOICE"
    elif has_sms and not has_voice:
        kind = "SMS"
    else:
        kind = None

    base = re.sub(r"\b(VOICE|SMS)\b", "", up).strip()
    return kind, base


def _phones_map_from_inv(inv: dict) -> dict[str, str]:
    phones: dict[str, str] = {}
    sp = inv.get("site_phones") or {}
    if isinstance(sp, dict):
        for raw_name, raw_val in sp.items():
            if not raw_name or raw_val is None:
                continue
            last4 = str(raw_val)[-4:]
            if not last4:
                continue
            phones[raw_name] = last4
            try:
                nk = _normalize_site_key(raw_name)
            except Exception:
                nk = None
            if nk:
                phones.setdefault(nk, last4)

    try:
        pm = _build_priority_phone_map(inv) or {}
        for raw_name, raw_val in pm.items():
            if not raw_name or raw_val is None:
                continue
            last4 = str(raw_val)[-4:]
            if not last4:
                continue
            phones.setdefault(raw_name, last4)
            try:
                nk = _normalize_site_key(raw_name)
                phones.setdefault(nk, last4)
            except Exception:
                pass
    except Exception:
        pass

    return phones


def decorate_with_last4_kind(inv: dict, desc: str) -> str:
    if not isinstance(desc, str) or not desc.strip():
        return desc

    if _re_dec.search(r"\(-\d{3,4}\)\s*$", desc):
        return desc

    kind, base = _infer_kind_and_base(desc)
    if not base:
        return desc

    if base.upper() in {"VOICE", "SMS"} and " " not in base.strip():
        return desc

    phones = _phones_map_from_inv(inv)

    candidates: list[str] = []
    candidates.append(desc)
    if kind:
        candidates.append(f"{base} {kind}")
    candidates.append(base)

    for key in candidates:
        last4 = phones.get(key)
        if last4:
            last4 = str(last4)[-4:]
            label = f"{base} {kind}" if kind else base
            return f"{label} (-{last4})"

    return desc


# ======================================================================
# QuickBooks export + division PDFs + template PDF
# (UNCHANGED from your version except for relying on fixed descriptions)
# ======================================================================

_QB_HEADER = [
    "*InvoiceNo",
    "*Customer",
    "*InvoiceDate",
    "*DueDate",
    "Terms",
    "Location",
    "Memo",
    "Item(Product/Service)",
    "ItemDescription",
    "ItemQuantity",
    "ItemRate",
    "*ItemAmount",
    "Service Date",
]


def _last_day_of_month(year: int, month: int) -> date:
    import calendar
    last = calendar.monthrange(year, month)[1]
    return date(year, month, last)


def _fmt_us_date(d: date) -> str:
    return f"{d.month}/{d.day}/{d.year}"


def _build_site_division_index_for_client(
    clients_doc: dict,
    client_name: str | None,
) -> tuple[list[str], dict[str, tuple[str, int, int]]]:
    if not client_name:
        return ([], {})

    client_name = client_name.strip()
    target = None
    for c in (clients_doc.get("clients") or []):
        if (c.get("name") or "").strip() == client_name:
            target = c
            break
    if not target:
        return ([], {})

    divisions = target.get("divisions") or []
    div_order: list[str] = []
    site_map: dict[str, tuple[str, int, int]] = {}

    for di, d in enumerate(divisions):
        dname = (d.get("name") or "").strip()
        if not dname:
            continue
        div_order.append(dname)
        for si, s in enumerate(d.get("sites") or []):
            sname = (s.get("name") or "").strip()
            if not sname:
                continue
            key = _normalize_site_key(sname)
            if key:
                site_map[key] = (dname, di, si)

    return (div_order, site_map)


def export_quickbooks_invoicing_csv(
    inv: Dict[str, Any],
    out_dir: str | Path | None = None,
    csv_name: str = "invoicing.csv",
) -> Path:
    period = inv.get("period") or {}
    year = int(period.get("year", 0) or 0)
    month = int(period.get("month", 0) or 0)
    if not year or not month:
        raise ValueError("Invoice 'period' is missing year/month.")

    client_name = inv.get("client_name_snapshot") or ""
    start_no = inv.get("starting_invoice_number")
    try:
        invoice_no = int(start_no)
    except Exception:
        invoice_no = 1

    invoice_date = _last_day_of_month(year, month)
    due_date = invoice_date
    service_date = date.today()

    inv_date_s = _fmt_us_date(invoice_date)
    due_date_s = _fmt_us_date(due_date)
    service_date_s = _fmt_us_date(service_date)

    clients_doc = _load_clients_doc()
    div_order, site_map = _build_site_division_index_for_client(
        clients_doc, client_name
    )

    by_div: dict[str, list[tuple[int, int, Dict[str, Any]]]] = {}
    leftovers: list[Dict[str, Any]] = []

    for li in inv.get("line_items", []):
        desc = str(li.get("description", "") or "")
        key = _normalize_site_key(desc)
        info = site_map.get(key)
        if not info:
            leftovers.append(li)
            continue
        dname, di, si = info
        by_div.setdefault(dname, []).append((di, si, li))

    rows: list[list[str]] = []

    def _add_invoice_for_division(
        div_name: str,
        items: list[tuple[int, int, Dict[str, Any]]],
        current_invoice_no: int,
    ) -> int:
        if not items:
            return current_invoice_no

        items_sorted = sorted(items, key=lambda t: t[1])

        first = True
        for _, _, li in items_sorted:
            qty_val = li.get("qty", 0)
            try:
                qty = str(int(round(float(qty_val))))
            except Exception:
                qty = str(qty_val)

            rate = float(li.get("unit_price", 0) or 0.0)
            amt = float(li.get("amount", 0) or 0.0)

            row: list[str] = []
            row.append(str(current_invoice_no))

            if first:
                row.extend([
                    client_name,
                    inv_date_s,
                    due_date_s,
                    "Due on Receipt",
                    "",
                    "",
                ])
                first = False
            else:
                row.extend(["", "", "", "", "", ""])

            row.extend([
                "Services",
                str(li.get("description", "")),
                qty,
                f"{rate:.2f}",
                f"{amt:.2f}",
                service_date_s,
            ])
            rows.append(row)

        return current_invoice_no + 1

    for dname in div_order:
        items = by_div.get(dname)
        if not items:
            continue
        invoice_no = _add_invoice_for_division(dname, items, invoice_no)

    for li in leftovers:
        invoice_no = _add_invoice_for_division(
            "(Unassigned)", [(0, 0, li)], invoice_no
        )

    out_dir = _ensure_out_dir_for_invoice(inv, out_dir)
    csv_path = out_dir / csv_name

    with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(_QB_HEADER)
        w.writerows(rows)

    return csv_path


def export_division_pdfs(
    inv: Dict[str, Any],
    template_path: str | Path | None = None,
    out_dir: str | Path | None = None,
) -> list[Path]:
    if out_dir is None:
        try:
            out_dir_path = _ensure_out_dir_for_invoice(inv, None)
        except Exception:
            out_dir_path = invoice_output_dir()
    else:
        out_dir_path = Path(out_dir)
        out_dir_path.mkdir(parents=True, exist_ok=True)

    client_name = inv.get("client_name_snapshot") or ""
    clients_doc = _load_clients_doc()
    div_order, site_map = _build_site_division_index_for_client(
        clients_doc, client_name
    )

    by_div: dict[str, list[Dict[str, Any]]] = {}
    leftovers: list[Dict[str, Any]] = []

    for li in inv.get("line_items", []):
        desc = str(li.get("description", "") or "")
        key = _normalize_site_key(desc)
        info = site_map.get(key)
        if not info:
            leftovers.append(li)
            continue
        dname, di, si = info
        by_div.setdefault(dname, []).append(li)

    results: list[Path] = []

    import copy, re as _re

    def _slugify(name: str) -> str:
        s = _re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-")
        return s or "div"

    base_id = str(inv.get("id", ""))

    tpl: Path | None = None
    if template_path is not None:
        p = Path(template_path)
        if p.exists() and os.name == "nt":
            tpl = p

    start_no = inv.get("starting_invoice_number")
    try:
        current_invoice_no = int(start_no)
    except Exception:
        current_invoice_no = 1

    def _make_pdf(div_inv: Dict[str, Any]) -> Path:
        if tpl is not None:
            path = export_invoice_pdf_via_template(
                div_inv, tpl, out_dir=out_dir_path
            )
            stem = invoice_filename(div_inv, "xlsm").replace(".xlsm", "")
            for ext in (".xlsm", ".xlsx", ".csv"):
                cand = out_dir_path / f"{stem}{ext}"
                if cand.exists():
                    try:
                        cand.unlink()
                    except Exception:
                        pass
            return Path(path)
        else:
            return export_invoice_pdf(div_inv, out_dir=out_dir_path)

    for dname in div_order:
        items = by_div.get(dname)
        if not items:
            continue
        div_inv = copy.deepcopy(inv)
        div_inv["division_name"] = dname
        div_inv["line_items"] = items
        div_inv["totals"] = {}
        div_inv["id"] = f"{base_id}-{_slugify(dname)}"
        div_inv["human_number"] = current_invoice_no
        recompute_totals(div_inv)
        pdf = _make_pdf(div_inv)
        results.append(pdf)
        current_invoice_no += 1

    if leftovers:
        dname = "(Unassigned)"
        div_inv = copy.deepcopy(inv)
        div_inv["division_name"] = dname
        div_inv["line_items"] = leftovers
        div_inv["totals"] = {}
        div_inv["id"] = f"{base_id}-{_slugify('unassigned')}"
        div_inv["human_number"] = current_invoice_no
        recompute_totals(div_inv)
        pdf = _make_pdf(div_inv)
        results.append(pdf)

    return results


def export_invoice_pdf_via_template(
    inv: Dict[str, Any],
    template_path: str | Path,
    out_dir: str | Path | None = None,
    clients_path: str | Path | None = None,
) -> Path:
    import openpyxl
    from openpyxl import load_workbook
    from datetime import datetime

    tpl = Path(template_path)
    if not tpl.exists():
        raise FileNotFoundError(f"Template not found: {tpl}")

    if out_dir is None:
        try:
            out_dir = invoice_output_dir()
        except Exception:
            out_dir = Path("invoices_output")
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(tpl, data_only=False, keep_vba=True)
    ws = wb.active

    invoice_number = (
        inv.get("human_number")
        or inv.get("qb_invoice_number")
        or inv.get("starting_invoice_number")
        or inv.get("id")
    )
    try:
        ws["G5"].value = str(invoice_number)
    except Exception:
        pass

    today = datetime.today()
    for cell in ("H5", "H7"):
        try:
            ws[cell].value = today
            ws[cell].number_format = "m/d/yyyy"
        except Exception:
            pass

    try:
        val = ws["H8"].value
        if val is None or str(val).strip() == "":
            ws["H8"].value = "Due on Receipt"
    except Exception:
        pass

    def _bill_to_lines() -> List[str]:
        try:
            doc = _load_clients_doc()
            client_name = (
                inv.get("client_name_snapshot")
                or inv.get("client_name")
                or ""
            )
            lines = _find_client_address(doc, client_name)
            if isinstance(lines, list) and lines:
                return [str(x) for x in lines][:3]
        except Exception:
            pass

        name = (
            inv.get("client_name_snapshot")
            or inv.get("client_name")
            or ""
        )
        addr = (
            inv.get("client_address_snapshot")
            or inv.get("client_address")
            or ""
        )

        out: List[str] = []
        if name:
            out.append(str(name))
        if addr:
            for line in str(addr).splitlines():
                line = line.strip()
                if line:
                    out.append(line)

        return out[:3]

    try:
        lines = _bill_to_lines()
        for i in range(3):
            ws[f"A{8 + i}"].value = lines[i] if i < len(lines) else None
    except Exception:
        pass

    row = 13
    line_items = inv.get("line_items", [])
    if not isinstance(line_items, list) or not line_items:
        line_items = inv.get("items", []) or []

    for li in line_items:
        try:
            raw_desc = li.get("description", "")
            desc = decorate_with_last4_kind(inv, raw_desc)
        except Exception:
            desc = li.get("description", "")

        try:
            ws[f"A{row}"].value = desc
            ws[f"F{row}"].value = float(li.get("qty", 0) or 0.0)
            ws[f"G{row}"].value = float(li.get("unit_price", 0) or 0.0)
            if ws[f"H{row}"].value in (None, ""):
                ws[f"H{row}"].value = f"=F{row}*G{row}"
        except Exception:
            pass
        row += 1

    last_item_row = max(13, row - 1)

    def _find_label_row(label: str) -> int | None:
        L = label.strip().upper()
        for r in range(13, 200):
            for c in range(1, 8):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip().upper() == L:
                    return r
        return None

    subtotal_row = _find_label_row("SUBTOTAL")
    total_row = _find_label_row("TOTAL")

    start_clear = last_item_row + 1
    if subtotal_row and start_clear < subtotal_row:
        for r in range(start_clear, subtotal_row):
            for col in ("A", "F", "G", "H"):
                try:
                    ws[f"{col}{r}"].value = None
                except Exception:
                    pass

    subtotal_formula = f"=SUM(H13:H{last_item_row})"
    try:
        if subtotal_row:
            cell = ws[f"H{subtotal_row}"]
            if cell.value in (None, ""):
                cell.value = subtotal_formula
        else:
            ws["H16"].value = subtotal_formula
            subtotal_row = 16
    except Exception:
        pass

    try:
        if total_row:
            cell = ws[f"H{total_row}"]
            if cell.value in (None, ""):
                cell.value = f"=H{subtotal_row}"
        else:
            ws["H17"].value = f"=H{subtotal_row}"
    except Exception:
        pass

    xlsm_path = out_dir / invoice_filename(inv, "xlsm")
    wb.save(xlsm_path)

    pdf_path = out_dir / invoice_filename(inv, "pdf")
    try:
        import win32com.client  # type: ignore
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        wb_com = excel.Workbooks.Open(str(xlsm_path))
        xlTypePDF = 0
        wb_com.ExportAsFixedFormat(xlTypePDF, str(pdf_path))
        wb_com.Close(False)
        excel.Quit()
        return pdf_path
    except Exception as e:
        raise RuntimeError(
            f"Excel export failed (install Excel + pywin32). Filled workbook at: {xlsm_path}"
        ) from e


# ======================================================================
# Message billing by segments (UPDATED to use make_site_description)
# ======================================================================

def _ceil_div2(n: int) -> int:
    try:
        x = int(float(n))
    except Exception:
        return 1
    if x <= 0:
        return 0
    return (x + 1) // 2

def _extract_num_segments(row: Dict[str, Any]) -> int:
    candidates = ("NumSegments", "Numsegments", "numsegments",
                  "Num_Segments", "NumSeg", "Numseg", "Segments", "segments")
    for k in candidates:
        if k in row and row[k] not in (None, "", "-"):
            try:
                return int(float(row[k]))
            except Exception:
                continue
    return 1

def _sum_billed_units_by_site(files_with_sites: List[Tuple[str | Path, str | None]],
                              year: int, month: int) -> Dict[str, int]:
    from collections import defaultdict
    totals: Dict[str, int] = defaultdict(int)

    for path, site_name in (files_with_sites or []):
        site = site_name or Path(path).stem
        try:
            with open(path, newline='', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        dt = _extract_row_datetime(row, "messages")
                    except Exception:
                        dt = None
                    if not dt or dt.year != int(year) or dt.month != int(month):
                        continue
                    seg = _extract_num_segments(row)
                    totals[site] += _ceil_div2(seg)
        except Exception:
            pass
    return dict(totals)

def add_message_items_to_invoice(
    inv: dict,
    messages_with_sites: List[Tuple[str | Path, str | None]],
    year: int,
    month: int,
    unit_price: float = 0.14,
) -> None:
    billed = _sum_billed_units_by_site(messages_with_sites, year, month)

    for site, qty in _ordered_site_items(billed):
        if qty <= 0:
            continue
        base = (site or "").strip()
        desc = make_site_description(base, "SMS") if base else "SMS"
        _add_item(inv, desc, qty, unit_price)

    recompute_totals(inv)


# ======================================================================
# Full CSV → PDF (unchanged)
# ======================================================================

def export_invoice_csv_full_pdf(csv_path: str | Path,
                                out_dir: str | Path | None = None) -> Path:
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
        from reportlab.lib import colors
    except Exception as e:
        raise RuntimeError(
            "reportlab is required to generate the full CSV table PDF "
            "(pip install reportlab)"
        ) from e

    csv_path = Path(csv_path)

    if out_dir is None:
        out_dir = csv_path.parent
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        rows = list(reader)

    if not rows:
        raise ValueError(f"CSV appears to be empty: {csv_path}")

    max_cols = max(len(r) for r in rows)
    norm_rows = [r + [""] * (max_cols - len(r)) for r in rows]

    pdf_path = out_dir / (csv_path.stem + "-full.pdf")

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=landscape(letter),
        leftMargin=24, rightMargin=24,
        topMargin=24, bottomMargin=24,
    )

    table = Table(norm_rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "LEFT"),
    ]))

    doc.build([table])
    return pdf_path
